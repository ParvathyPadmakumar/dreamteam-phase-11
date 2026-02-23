import os 
import openpyxl 
import csv
for excel_file in os.listdir('.'):
    if not excel_file.endswith('.xlsx'):
        continue
    wb=openpyxl.load_workbook(excel_file)
  
    for sheet_name in wb.sheetnames:
        sheet=wb[sheet_name]
        csvfile=f'{os.path.splitext(excel_file)[0]}_{sheet_name}.csv'    
        print("csv file:",csvfile)
        writer=csv.writer(open(csvfile,'w',newline='')) 
        for row_num in range(1,sheet.max_row+1):
            rows=[]   
            for col_num in range(1, sheet.max_column+1):
                cell=sheet.cell(row=row_num,column=col_num).value
                rows.append(cell)
            writer.writerow(rows)
